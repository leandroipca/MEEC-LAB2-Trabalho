import cv2
import os
import numpy as np
from openpyxl import Workbook
arquivo_excel = Workbook()
from openpyxl import load_workbook
arquivo_excel = load_workbook('resultados.xlsx')
from PIL import Image

detectorFace = cv2.CascadeClassifier("Haar/haarcascade_frontalface_default.xml")

#Tipos de classificador

reconhecedorEingen = cv2.face.EigenFaceRecognizer_create()
reconhecedorEingen.read("classificadorEigen_Teste.yml")
reconhecedorFisher = cv2.face.FisherFaceRecognizer_create()
reconhecedorFisher.read("classificadorFisher_Teste.yml")
reconhecedorLBPH = cv2.face.LBPHFaceRecognizer_create()
reconhecedorLBPH.read("classificadorLBPH_Teste.yml")

totalAcertosEigen = totalAcertosFisher = totalAcertosLBPH  = 0
percentualAcertoEigen = percentualAcertoFisher = percentualAcertoLBPH = 0.0
totalConfiancaEigen = totalConfiancaFisher = totalConfiancaLBPH  = 0.0



caminhos = [os.path.join('testes_reconhecimento/teste', f) for f in os.listdir('testes_reconhecimento/teste')]

# Classificador Eigen

for caminhoImagem in caminhos:
    imagemFace = Image.open(caminhoImagem).convert('L')
    imagemFaceNP = np.array(imagemFace, 'uint8')
    facesDetectadas = detectorFace.detectMultiScale(imagemFaceNP)
    for (x, y, l, a) in facesDetectadas:
        idprevistoEigen, confiancaEigen = reconhecedorEingen.predict(imagemFaceNP)
        idatual = int(os.path.split(caminhoImagem)[1].split('_')[0].replace("G", ""))
        print(str(idatual) + " foi classificado como " + str(idprevistoEigen) + " - " + str(confiancaEigen))
        if idprevistoEigen == idatual:
            totalAcertosEigen += 1
            totalConfiancaEigen += confiancaEigen
        cv2.rectangle(imagemFaceNP, (x, y), (x + l, y + a), (0, 0, 255), 2)
        #cv2.imshow("Eigen", imagemFaceNP)
        cv2.waitKey(1000)
percentualAcertoEigen = (totalAcertosEigen / 8) * 100
totalConfiancaEigen = totalConfiancaEigen / totalAcertosEigen
print("Percentual de acerto: " + str(percentualAcertoEigen) + "%")
print("Total confiança: " + str(totalConfiancaEigen))


# Classificador Fisher

for caminhoImagem in caminhos:
    imagemFace = Image.open(caminhoImagem).convert('L')
    imagemFaceNP = np.array(imagemFace, 'uint8')
    facesDetectadas = detectorFace.detectMultiScale(imagemFaceNP)
    for (x, y, l, a) in facesDetectadas:
        idprevistoFisher, confiancaFisher = reconhecedorFisher.predict(imagemFaceNP)
        idatual = int(os.path.split(caminhoImagem)[1].split('_')[0].replace("G", ""))
        print(str(idatual) + " foi classificado como " + str(idprevistoFisher) + " - " + str(confiancaFisher))
        if idprevistoFisher == idatual:
            totalAcertosFisher += 1
            totalConfiancaFisher += confiancaFisher
        cv2.rectangle(imagemFaceNP, (x, y), (x + l, y + a), (0, 0, 255), 2)
        #cv2.imshow("Fisher", imagemFaceNP)
        cv2.waitKey(1000)
percentualAcertoFisher = (totalAcertosFisher / 8) * 100
totalConfiancaFisher = totalConfiancaFisher / totalAcertosFisher
print("Percentual de acerto: " + str(percentualAcertoFisher) + "%")
print("Total confiança: " + str(totalConfiancaFisher))

# Classificador LBPH

for caminhoImagem in caminhos:
    imagemFace = Image.open(caminhoImagem).convert('L')
    imagemFaceNP = np.array(imagemFace, 'uint8')
    facesDetectadas = detectorFace.detectMultiScale(imagemFaceNP)
    for (x, y, l, a) in facesDetectadas:
        idprevistoLBPH, confiancLBPH = reconhecedorLBPH.predict(imagemFaceNP)
        idatual = int(os.path.split(caminhoImagem)[1].split('_')[0].replace("G", ""))
        print(str(idatual) + " foi classificado como " + str(idprevistoLBPH) + " - " + str(totalConfiancaLBPH))
        if idprevistoLBPH == idatual:
            totalAcertosLBPH += 1
            totalConfiancaLBPH += confiancLBPH
        cv2.rectangle(imagemFaceNP, (x, y), (x + l, y + a), (0, 0, 255), 2)
        #cv2.imshow("LBPH", imagemFaceNP)
        cv2.waitKey(1000)
percentualAcertoLBPH = (totalAcertosLBPH / 8) * 100
totalConfiancaLBPH = totalConfiancaLBPH / totalAcertosLBPH
print("Percentual de acerto: " + str(percentualAcertoLBPH) + "%")
print("Total confiança: " + str(totalConfiancaLBPH))


planilha1 = arquivo_excel.active
planilha1.title = "Comparador"
planilha1['A3'] = 'Eigen'
planilha1['A4'] = 'Acertos'
planilha1['B4'] = 'Porcentagem de Acertos'
planilha1['C4'] = 'Confiança'
planilha1['A5'] = totalAcertosEigen
planilha1['B5'] = str((round(percentualAcertoEigen)))
planilha1['C5'] = str((round(totalConfiancaEigen)))
planilha1['E3'] = 'Fisher'
planilha1['E4'] = 'Acertos'
planilha1['F4'] = 'Porcentagem de Acertos'
planilha1['G4'] = 'Confiança'
planilha1['E5'] = totalAcertosFisher
planilha1['F5'] = str((round(percentualAcertoFisher)))
planilha1['G5'] = str((round(totalConfiancaFisher)))
planilha1['I3'] = 'LBPH'
planilha1['I4'] = 'Acertos'
planilha1['J4'] = 'Porcentagem de Acertos'
planilha1['K4'] = 'Confiança'
planilha1['I5'] = totalAcertosLBPH
planilha1['J5'] = str((round(percentualAcertoLBPH)))
planilha1['K5'] = str((round(totalConfiancaLBPH)))
arquivo_excel.save("resultados.xlsx")





